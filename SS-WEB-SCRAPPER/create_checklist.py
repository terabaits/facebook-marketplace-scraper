import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Arsta Birojs Learning Checklist"

# Define styles
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
category_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Set column widths
ws.column_dimensions['A'].width = 8   # Checkbox column
ws.column_dimensions['B'].width = 70  # Task column
ws.column_dimensions['C'].width = 35  # Category column
ws.column_dimensions['D'].width = 18  # Status dropdown column
ws.column_dimensions['E'].width = 40  # Notes column
ws.column_dimensions['F'].width = 15  # Date completed

# Title row
ws.merge_cells('A1:F1')
ws['A1'] = "ARSTA BIROJS v46 - SISTEMAS LIETOTAJU ROKASGRAMATA - MACIBU UZDEVUMU SARAKSTS"
ws['A1'].font = Font(bold=True, size=14)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 30

# Header row
headers = ['☐', 'Uzdevums / Macibu vieniba', 'Kategorija', 'Statuss', 'Piezimes', 'Pabeigts']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border
ws.row_dimensions[2].height = 25

# Define all tasks/learning items from the document
tasks = [
    # 1. Ievads un pamati
    ("Ievads un sistēmas pamati", [
        ("Iepazities ar sistemas Arsta Birojs v5 darba uzsaksanu", "Ievads un pamati"),
        ("Iepazities ar sistemas Arsta Birojs v5 darba beigsanu", "Ievads un pamati"),
        ("Izmantot sistemas galveno skatu un izvelni", "Ievads un pamati"),
        ("Macities stradat ar sistemas logiem un sarakstiem", "Ievads un pamati"),
        ("Apgut sarakstu kartosanu un meklesanu", "Ievads un pamati"),
        ("Apgut datu filtresanas iespejas", "Ievads un pamati"),
        ("Macities lietot kursora uzvedibas rezimu", "Ievads un pamati"),
        ("Iemacities sarakstu eksportesanu un drukasanu", "Ievads un pamati"),
        ("Apgut rindas iesaldesanu (freeze panes)", "Ievads un pamati"),
        ("Iemacities pievienot un labot ierakstus", "Ievads un pamati"),
        ("Iepazities ar sistemas zinojumiem un pieteikumu registraciju", "Ievads un pamati"),
    ]),
    
    # 2. Lietotaju parvaldiba
    ("Lietotaju parvaldiba", [
        ("Jauna lietotaja ievade sistema", "Lietotaju parvaldiba"),
        ("Lietotaja pamatdatu maina", "Lietotaju parvaldiba"),
        ("Paroles maina un drosibas iestatijumi", "Lietotaju parvaldiba"),
        ("Lietotaju tiesibu parvaldiba", "Lietotaju parvaldiba"),
        ("Uzdevumu rindas izmantosana", "Lietotaju parvaldiba"),
        ("Lietotaju grupu konfiguresana", "Lietotaju parvaldiba"),
    ]),
    
    # 3. Pamatklasifikatori
    ("Pamatklasifikatori", [
        ("Iestazu klasifikatora uzturesana", "Pamatklasifikatori"),
        ("Strukturvienibu klasifikatora uzturesana", "Pamatklasifikatori"),
        ("Specialitasu klasifikatora uzturesana", "Pamatklasifikatori"),
        ("Ligumattiecibu klasifikatora uzturesana", "Pamatklasifikatori"),
        ("Diagnosu klasifikatora (SSK-10) uzturesana", "Pamatklasifikatori"),
        ("Manipulaciju klasifikatora uzturesana", "Pamatklasifikatori"),
        ("Manipulaciju grupu klasifikatora uzturesana", "Pamatklasifikatori"),
        ("Manipulaciju paku klasifikatora uzturesana", "Pamatklasifikatori"),
        ("Operaciju klasifikatora uzturesana", "Pamatklasifikatori"),
        ("Administrativo teritoriju (ATVK) klasifikatora uzturesana", "Pamatklasifikatori"),
        ("Pacientu grupu klasifikatora uzturesana", "Pamatklasifikatori"),
        ("E veidlapu izsniedzeju klasifikatora uzturesana", "Pamatklasifikatori"),
        ("E veidlapu klasifikatora uzturesana", "Pamatklasifikatori"),
        ("Svitrkodu klasifikatora uzturesana", "Pamatklasifikatori"),
    ]),
    
    # 4. Pacienta klasifikatori
    ("Pacienta klasifikatori", [
        ("Tautibu klasifikatora uzturesana", "Pacienta klasifikatori"),
        ("Gimenes stavoklu klasifikatora uzturesana", "Pacienta klasifikatori"),
        ("Ambulatoro karšu statusu klasifikatora uzturesana", "Pacienta klasifikatori"),
        ("Informacijas avotu klasifikatora uzturesana", "Pacienta klasifikatori"),
        ("Alergiju klasifikatora uzturesana", "Pacienta klasifikatori"),
        ("Anamnezes jautajumu klasifikatora uzturesana", "Pacienta klasifikatori"),
        ("Asins grupu klasifikatora uzturesana", "Pacienta klasifikatori"),
        ("Svariguma pakapju klasifikatora uzturesana", "Pacienta klasifikatori"),
        ("Profilaksu/pasakumu veidu klasifikatora uzturesana", "Pacienta klasifikatori"),
        ("Valsts piederibas klasifikatora uzturesana", "Pacienta klasifikatori"),
        ("Invaliditates grupu klasifikatora uzturesana", "Pacienta klasifikatori"),
        ("Invaliditates veidu klasifikatora uzturesana", "Pacienta klasifikatori"),
        ("Seksualas orientacijas klasifikatora uzturesana", "Pacienta klasifikatori"),
        ("Izglitibas limena klasifikatora uzturesana", "Pacienta klasifikatori"),
        ("Saites veidu klasifikatora uzturesana", "Pacienta klasifikatori"),
        ("Numeresanas masku konfiguresana", "Pacienta klasifikatori"),
    ]),
    
    # 5. Kalendara klasifikatori
    ("Kalendara klasifikatori", [
        ("Istabu klasifikatora uzturesana", "Kalendara klasifikatori"),
        ("Arstu un istabu atbilstibas klasifikatora uzturesana", "Kalendara klasifikatori"),
        ("Strukturvienibu un istabu atbilstibas klasifikatora uzturesana", "Kalendara klasifikatori"),
        ("Kalendara pieraksta veidu klasifikatora uzturesana", "Kalendara klasifikatori"),
        ("Darba laiku veidu klasifikatora uzturesana", "Kalendara klasifikatori"),
        ("Kalendara pieraksta krasu klasifikatora uzturesana", "Kalendara klasifikatori"),
        ("Kalendara pieraksta iconu klasifikatora uzturesana", "Kalendara klasifikatori"),
        ("Brivdienu klasifikatora uzturesana", "Kalendara klasifikatori"),
        ("Proceduru klasifikatora uzturesana", "Kalendara klasifikatori"),
    ]),
    
    # 6. Ambulatorie klasifikatori
    ("Ambulatorie klasifikatori", [
        ("Specialistiem atlauto manipulaciju klasifikatora uzturesana", "Ambulatorie klasifikatori"),
        ("Maksataju (ambulatorajiem taloniem) klasifikatora uzturesana", "Ambulatorie klasifikatori"),
        ("Testa rezultatu kodu klasifikatora uzturesana", "Ambulatorie klasifikatori"),
        ("Diagnosu un manipulaciju atbilstibas klasifikatora uzturesana", "Ambulatorie klasifikatori"),
        ("Strukturvienibu manipulaciju klasifikatora uzturesana", "Ambulatorie klasifikatori"),
    ]),
    
    # 7. Stacionarie klasifikatori
    ("Stacionarie klasifikatori", [
        ("Traumu veidu klasifikatora uzturesana", "Stacionarie klasifikatori"),
        ("Gultu profilu klasifikatora uzturesana", "Stacionarie klasifikatori"),
        ("Nodalju strukturvienibu klasifikatora uzturesana", "Stacionarie klasifikatori"),
        ("Nodalju klasifikatora uzturesana (ar palatam un gultam)", "Stacionarie klasifikatori"),
        ("Kustibu veidu klasifikatora uzturesana", "Stacionarie klasifikatori"),
        ("Ilgtermina rindu statusu klasifikatora uzturesana", "Stacionarie klasifikatori"),
        ("Ilgtermina rindu veidu klasifikatora uzturesana", "Stacionarie klasifikatori"),
        ("Stacionaro karšu statusu klasifikatora uzturesana", "Stacionarie klasifikatori"),
        ("Vestures ieraksta tipu klasifikatora uzturesana", "Stacionarie klasifikatori"),
        ("Palatu kategoriju klasifikatora uzturesana", "Stacionarie klasifikatori"),
        ("Uzņemšanas prioritāšu klasifikatora uzturesana", "Stacionarie klasifikatori"),
        ("Manipulaciju paku stacionaram klasifikatora uzturesana", "Stacionarie klasifikatori"),
    ]),
    
    # 8. Vizualas diagnostikas klasifikatori
    ("Vizualas diagnostikas klasifikatori", [
        ("Ķermenja dalu klasifikatora uzturesana", "Vizualas diagnostikas klasifikatori"),
        ("Nosutijumu veidu klasifikatora uzturesana", "Vizualas diagnostikas klasifikatori"),
        ("Filmu klasifikatora uzturesana", "Vizualas diagnostikas klasifikatori"),
        ("Filmu dalu klasifikatora uzturesana", "Vizualas diagnostikas klasifikatori"),
        ("Poziciju klasifikatora uzturesana", "Vizualas diagnostikas klasifikatori"),
        ("Projekciju klasifikatora uzturesana", "Vizualas diagnostikas klasifikatori"),
    ]),
    
    # 9. Pakalpojumu klasifikatori
    ("Pakalpojumu klasifikatori", [
        ("Maksas pakalpojumu klasifikatora uzturesana", "Pakalpojumu klasifikatori"),
        ("Pakalpojumu sadalijuma klasifikatora uzturesana", "Pakalpojumu klasifikatori"),
        ("NCSP kodu klasifikatora uzturesana", "Pakalpojumu klasifikatori"),
        ("Pakalpojumu un manipulaciju sasaites klasifikatora uzturesana", "Pakalpojumu klasifikatori"),
    ]),
    
    # 10. Kases klasifikatori
    ("Kases klasifikatori", [
        ("Kases klasifikatora uzturesana", "Kases klasifikatori"),
        ("Apmaksu tipu klasifikatora uzturesana", "Kases klasifikatori"),
        ("Davanu karšu pakalpojumu kodu iestatīsana", "Kases klasifikatori"),
        ("OVP pakalpojumu klasifikatora uzturesana", "Kases klasifikatori"),
    ]),
    
    # 11. Iestatijumi
    ("Iestatijumu parvaldiba", [
        ("Globalo iestatijumu konfiguresana", "Iestatijumi"),
        ("Lietotaja iestatijumu parvaldiba", "Iestatijumi"),
        ("Ambulatora talona iestatijumu konfiguresana", "Iestatijumi"),
        ("Stacionaras kartes iestatijumu konfiguresana", "Iestatijumi"),
        ("Nosutijuma iestatijumu konfiguresana", "Iestatijumi"),
        ("Ziņapmaijnas iestatijumu konfiguresana", "Iestatijumi"),
        ("Asins kabineta iestatijumu konfiguresana", "Iestatijumi"),
        ("Rindas iestatijumu konfiguresana", "Iestatijumi"),
        ("Recepšu iestatijumu konfiguresana", "Iestatijumi"),
        ("Zobarstniecibas talona iestatijumu konfiguresana", "Iestatijumi"),
        ("Uzņemšanas/traumpunkta iestatijumu konfiguresana", "Iestatijumi"),
        ("NMC iestatijumu konfiguresana", "Iestatijumi"),
        ("Tafeles iestatijumu konfiguresana", "Iestatijumi"),
        ("Konfiguracijas parametru iestatīsana", "Iestatijumi"),
        ("Komplekso kontrolu iestatijumu konfiguresana", "Iestatijumi"),
        ("Bridinajumu iestatijumu konfiguresana", "Iestatijumi"),
    ]),
    
    # 12. Darbs ar pacientiem
    ("Darbs ar pacientiem", [
        ("Jauna pacienta registrēsana", "Darbs ar pacientiem"),
        ("Pacienta pamatdatu ievade un labosana", "Darbs ar pacientiem"),
        ("Ambulatoras kartes parvaldiba", "Darbs ar pacientiem"),
        ("Stacionaras kartes parvaldiba", "Darbs ar pacientiem"),
        ("Pacientu grupu pieskirsana", "Darbs ar pacientiem"),
        ("Diagnosu ievade pacientam", "Darbs ar pacientiem"),
        ("Alergiju ievade pacientam", "Darbs ar pacientiem"),
        ("Dzives anamnezes jautajumu parvaldiba", "Darbs ar pacientiem"),
        ("Asins grupas un parliesanas datu parvaldiba", "Darbs ar pacientiem"),
        ("Profilakses un pasakumu parvaldiba", "Darbs ar pacientiem"),
        ("Invaliditates datu parvaldiba", "Darbs ar pacientiem"),
        ("Pacientu saistibu parvaldiba", "Darbs ar pacientiem"),
        ("Primaras aprupes datu parvaldiba", "Darbs ar pacientiem"),
        ("Apdrosinasanas datu parvaldiba", "Darbs ar pacientiem"),
        ("Failu parvaldiba pacienta karte", "Darbs ar pacientiem"),
        ("Pacientu meklesana un atlase", "Darbs ar pacientiem"),
        ("Mani pacienti funkcionalitates izmantosana", "Darbs ar pacientiem"),
        ("Pazimju un markejumu parvaldiba", "Darbs ar pacientiem"),
    ]),
    
    # 13. Ambulatorais talons
    ("Ambulatorais talons", [
        ("Jauna ambulatora talona izveidosana", "Ambulatorais talons"),
        ("Ambulatora talona meklesana", "Ambulatorais talons"),
        ("Talona datu ievade (diagnoses, manipulacijas)", "Ambulatorais talons"),
        ("Talona noslegsana", "Ambulatorais talons"),
        ("Talona anulešana", "Ambulatorais talons"),
        ("Atkartotas talona veidosana 30 dienu perioda", "Ambulatorais talons"),
        ("Apmeklejumu parvaldiba talona", "Ambulatorais talons"),
        ("Manipulaciju pievienošana talona", "Ambulatorais talons"),
        ("Diagnosu un manipulaciju automatiska ielade", "Ambulatorais talons"),
        ("NCSP izmeklejumu parvaldiba talona", "Ambulatorais talons"),
        ("Primaro datu parvaldiba talona", "Ambulatorais talons"),
    ]),
    
    # 14. Stacionars
    ("Stacionara nodala", [
        ("Pacienta uzņemšana nodala", "Stacionara nodala"),
        ("Stacionara kartes meklesana", "Stacionara nodala"),
        ("Stacionara kartes veidosana", "Stacionara nodala"),
        ("Kustibu veidu parvaldiba", "Stacionara nodala"),
        ("Nodaljas maina", "Stacionara nodala"),
        ("Gultas profila maina", "Stacionara nodala"),
        ("Pacienta parvietošana starp nodalam", "Stacionara nodala"),
        ("Pacienta izrakstišana", "Stacionara nodala"),
        ("Stacionara karšu slegsana", "Stacionara nodala"),
        ("Izraksta-epikrizes parvaldiba", "Stacionara nodala"),
        ("Vestures ierakstu parvaldiba", "Stacionara nodala"),
    ]),
    
    # 15. Uzņemšana un traumpunkts
    ("Uzņemšana un traumpunkts", [
        ("Uzņemšanas/traumpunkta ierakstu veidosana", "Uzņemšana un traumpunkts"),
        ("Uzņemšanas rindas meklesana", "Uzņemšana un traumpunkts"),
        ("Prioritates pieskirsana", "Uzņemšana un traumpunkts"),
        ("Nodaljas izvele un maina", "Uzņemšana un traumpunkts"),
        ("Hospitalizacijas planošana", "Uzņemšana un traumpunkts"),
        ("Ambulatora talona generēsana no uzņemšanas", "Uzņemšana un traumpunkts"),
        ("NMC (Neatliekamas mediciniskas palidzibas) parvaldiba", "Uzņemšana un traumpunkts"),
        ("Pacientu Melnais saraksts", "Uzņemšana un traumpunkts"),
    ]),
    
    # 16. Nosutijumi (RIS)
    ("Nosutijumi (RIS)", [
        ("Jauna nosutijuma izveidosana", "Nosutijumi"),
        ("Nosutijumu meklesana", "Nosutijumi"),
        ("Izmeklejuma veikšana", "Nosutijumi"),
        ("Sledzienu rakstišana", "Nosutijumi"),
        ("Nosutijuma blokešana un atblokešana", "Nosutijumi"),
        ("Ambulatora talona generēsana no nosutijuma", "Nosutijumi"),
        ("Bilžu apskatišana (Jivex, IO-VIEW, ClearCanvas)", "Nosutijumi"),
        ("DataMed integracija", "Nosutijumi"),
    ]),
    
    # 17. Kalendars
    ("Kalendars un pieraksti", [
        ("Darba laiku definesana", "Kalendars un pieraksti"),
        ("Darba laiku meklesana", "Kalendars un pieraksti"),
        ("Darba laiku masveida labošana", "Kalendars un pieraksti"),
        ("Kalendara pierakstu izveidosana", "Kalendars un pieraksti"),
        ("Kalendara pieraksta parcelesana un kopešana", "Kalendars un pieraksti"),
        ("Attalinatas konsultacijas pieraksts", "Kalendars un pieraksti"),
        ("Kalendara papildinformacijas parvaldiba", "Kalendars un pieraksti"),
        ("Arejo pierakstu parvaldiba", "Kalendars un pieraksti"),
        ("Pieraksta vestures apskate", "Kalendars un pieraksti"),
        ("Apmeklejumu parskats", "Kalendars un pieraksti"),
        ("Arstu prombutņu parvaldiba", "Kalendars un pieraksti"),
    ]),
    
    # 18. Planošanas kalendars
    ("Planošanas kalendars", [
        ("Apmeklejuma planošana", "Planošanas kalendars"),
        ("Pieraksta kopešana perioda", "Planošanas kalendars"),
        ("Ieradušos ķeķša kopešana", "Planošanas kalendars"),
        ("Kopsavilkuma atskaites veidosana", "Planošanas kalendars"),
        ("Planošanas kalendara iestatijumi", "Planošanas kalendars"),
    ]),
    
    # 19. Kase un aprekini
    ("Kase un aprekini", [
        ("Aprekina izveidosana", "Kase un aprekini"),
        ("Aprekina meklesana", "Kase un aprekini"),
        ("Reķina izveidosana", "Kase un aprekini"),
        ("Reķina apmaksa", "Kase un aprekini"),
        ("Reķina anulešana", "Kase un aprekini"),
        ("Kases ķeku parvaldiba", "Kase un aprekini"),
        ("Darbs ar kases ķeku meklesanas logu", "Kase un aprekini"),
        ("Spogula ekrana izmantosana", "Kase un aprekini"),
        ("Ilgtermina rindas", "Kase un aprekini"),
        ("Davanu karšu parvaldiba", "Kase un aprekini"),
        ("Atlaižu karšu parvaldiba", "Kase un aprekini"),
        ("Maksas pakalpojumu parvaldiba", "Kase un aprekini"),
        ("Aprekina parskata izdruka", "Kase un aprekini"),
        ("Reķinu izdruka", "Kase un aprekini"),
    ]),
    
    # 20. E-Veselibas dokumenti
    ("E-Veselibas dokumenti", [
        ("Darbs ar E-Veselibas dokumentiem", "E-Veselibas dokumenti"),
        ("Dokumentu meklesanas panelis", "E-Veselibas dokumenti"),
        ("Dokumentu saraksta parvaldiba", "E-Veselibas dokumenti"),
        ("Darbnespejas lapas (DNL) forma", "E-Veselibas dokumenti"),
        ("DNL izrakstišana un slegsana", "E-Veselibas dokumenti"),
        ("DNL anulešana un kopešana", "E-Veselibas dokumenti"),
        ("DNL parņemšana", "E-Veselibas dokumenti"),
        ("DNL deaktivizacija E-veseliba", "E-Veselibas dokumenti"),
        ("Recepšu forma un drukasana", "E-Veselibas dokumenti"),
        ("Recepšu kopešana", "E-Veselibas dokumenti"),
        ("Autotekstu izmantosana receptes", "E-Veselibas dokumenti"),
        ("Dokumentu izgušana no E-Veselibas", "E-Veselibas dokumenti"),
        ("Dokumentu apstiprinašana un nosutišana", "E-Veselibas dokumenti"),
        ("Dokumentu drukasana", "E-Veselibas dokumenti"),
        ("Ambulatora nosutijuma dokumenti", "E-Veselibas dokumenti"),
        ("Radiologiska izmeklejuma apraksts", "E-Veselibas dokumenti"),
        ("Ambulatoras vizites parskats", "E-Veselibas dokumenti"),
        ("Izraksts-epikrize", "E-Veselibas dokumenti"),
        ("Pacienta veselibas kartes papildinajums (PVK)", "E-Veselibas dokumenti"),
    ]),
    
    # 21. OVP lapas
    ("OVP lapas", [
        ("OVP lapas izveidosana", "OVP lapas"),
        ("Veselibai kaitigo darba vides faktoru noradišana", "OVP lapas"),
        ("Darbu ipašos apstaklos noradišana", "OVP lapas"),
        ("Specialitasu un izmeklejumu saraksta generēsana", "OVP lapas"),
    ]),
    
    # 22. Profilakse un pasakumi
    ("Profilakse un pasakumi", [
        ("Profilakses/pasakumu pievienošana", "Profilakse un pasakumi"),
        ("Profilakses pasakumu planošana", "Profilakse un pasakumi"),
        ("MediCloud profilakses pasakumi", "Profilakse un pasakumi"),
        ("Vakcinacijas parvaldiba", "Profilakse un pasakumi"),
        ("Potes registrēsana", "Profilakse un pasakumi"),
    ]),
    
    # 23. Laboratorija
    ("Laboratorija", [
        ("Laboratorijas nosutijumu izveide", "Laboratorija"),
        ("Laboratorijas iestatijumu konfiguresana", "Laboratorija"),
        ("Registracijas zurnala parvaldiba", "Laboratorija"),
        ("Testešanas protokolu parvaldiba", "Laboratorija"),
        ("Iekartu zurnalu parvaldiba", "Laboratorija"),
        ("Paraugu zurnalu parvaldiba", "Laboratorija"),
        ("Dialab integracija", "Laboratorija"),
    ]),
    
    # 24. Apdrosinasana
    ("Apdrosinasana", [
        ("Apdrosinataju iestatijumu konfiguresana (HWS, ERGO, BTA, Seesam, Balta, Compensa, BAN, Gjensidige)", "Apdrosinasana"),
        ("Polišu un maksas pakalpojumu sasait", "Apdrosinasana"),
        ("Apdrosinasanas datu parvaldiba pacienta karte", "Apdrosinasana"),
        ("QR kodu izmantosana apdrosinasana", "Apdrosinasana"),
    ]),
    
    # 25. Tafele
    ("Nodaljas tafele", [
        ("Tafeles skata izmantosana", "Nodaljas tafele"),
        ("Lietotaja darbibas tafele", "Nodaljas tafele"),
        ("Gultu pieskirsana pacientam", "Nodaljas tafele"),
        ("Masu un aprupes limenu pieskirsana", "Nodaljas tafele"),
        ("Dietu pieskirsana", "Nodaljas tafele"),
        ("Ipašo dietu parvaldiba", "Nodaljas tafele"),
        ("Paraugu noņemšanas pazimju parvaldiba", "Nodaljas tafele"),
        ("Palatu un gultu kartošana", "Nodaljas tafele"),
    ]),
    
    # 26. Ziņapmaijna
    ("Ziņapmaijna", [
        ("Zinojumu sutišana un sanemšana", "Ziņapmaijna"),
        ("Zinojumu parvaldiba", "Ziņapmaijna"),
        ("Saziņa ar pacientu", "Ziņapmaijna"),
        ("Reklamu un piedavajumu parvaldiba", "Ziņapmaijna"),
    ]),
    
    # 27. Atskaites un eksports
    ("Atskaites un eksports", [
        ("Datu eksports", "Atskaites un eksports"),
        ("Dokumentu eksports uz arejam sistemam", "Atskaites un eksports"),
        ("Atskaišu papildus rekvizitu konfiguresana", "Atskaites un eksports"),
        ("Izmantošana divas datu bazes atskaišu veidosanai", "Atskaites un eksports"),
    ]),
    
    # 28. Nieru transplantacija
    ("Nieru transplantacijas pacientu registrs", [
        ("Nieru transplantacijas pacientu registra parvaldiba", "Nieru transplantacija"),
    ]),
    
    # 29. MediCloud integracija
    ("MediCloud integracija", [
        ("MediCloud iestatijumu konfiguresana", "MediCloud"),
        ("MediCloud pakalpojumu parvaldiba", "MediCloud"),
        ("Lojalitates programmas apmaijna", "MediCloud"),
        ("Vakcinacijas datu apmaijna", "MediCloud"),
        ("Globalas ziņapmaijnas izmantosana", "MediCloud"),
    ]),
    
    # 30. Q-Flow integracija
    ("Q-Flow integracija", [
        ("AB integracija ar rindu sistemu Q-Flow", "Q-Flow"),
        ("Q-Flow servisa adreses uzstadīsana", "Q-Flow"),
    ]),
    
    # 31. Pašapkalposanas stends
    ("Pašapkalposanas stends", [
        ("Pašapkalposanas stenda iestatijumu konfiguresana", "Pašapkalposanas stends"),
    ]),
    
    # 32. Jaundzimusie
    ("Jaundzimuso parvaldiba", [
        ("Jaundzimusa registrēsana", "Jaundzimuso parvaldiba"),
        ("Datu kopešana no mates kartes", "Jaundzimuso parvaldiba"),
    ]),
    
    # 33. Zobarstnieciba
    ("Zobarstniecibas talons", [
        ("Zobarstniecibas talona izveidosana", "Zobarstniecibas talons"),
        ("Zobarstniecibas talona iestatijumu konfiguresana", "Zobarstniecibas talons"),
    ]),
    
    # 34. Specialie rezimi
    ("Specialie funkciju rezimi", [
        ("Testa rezima izmantosana", "Specialie rezimi"),
        ("Filiāļu rezima izmantosana", "Specialie rezimi"),
        ("IMPAX zinojumu sutišana", "Specialie rezimi"),
        ("Auditacijas atsleas izmantosana", "Specialie rezimi"),
    ]),
    
    # 35. Reķinu parvaldiba
    ("Reķinu parvaldiba", [
        ("Reķinu numerešanas iestatijumi", "Reķinu parvaldiba"),
        ("Nodalju lietošana reķinos", "Reķinu parvaldiba"),
        ("Apmaksas terminu iestatīsana", "Reķinu parvaldiba"),
        ("Valutu izveles konfiguresana", "Reķinu parvaldiba"),
        ("Reķinu izdrukas iestatijumi", "Reķinu parvaldiba"),
        ("Logotipa pievienošana reķiniem", "Reķinu parvaldiba"),
    ]),
    
    # 36. ķeka un kases parvaldiba
    ("ķeka un kases parvaldiba", [
        ("ķeka iestatijumu konfiguresana", "ķeka un kases parvaldiba"),
        ("Kases parametru iestatīsana", "ķeka un kases parvaldiba"),
        ("ķeku lentes platuma iestatīsana", "ķeka un kases parvaldiba"),
        ("Kases draivera konfiguresana", "ķeka un kases parvaldiba"),
        ("Kases servera pieslegums", "ķeka un kases parvaldiba"),
        ("ķeka izdrukas papildinajumi", "ķeka un kases parvaldiba"),
        ("Horizon importeto maksajumu kases iestatīsana", "ķeka un kases parvaldiba"),
    ]),
    
    # 37. Horizon integracija
    ("Horizon integracija", [
        ("Horizon iestatijumu konfiguresana", "Horizon"),
        ("Horizon sinhronizacija", "Horizon"),
        ("Reķinu, apmaksu un anulaciju sinhronizacija", "Horizon"),
        ("Horizon REST API izmantosana", "Horizon"),
    ]),
    
    # 38. Datu apmaijna
    ("Datu apmaijna", [
        ("Datu apmaijnas moduļa izmantosana", "Datu apmaijna"),
        ("Failu importa/eksporta parvaldiba", "Datu apmaijna"),
        ("NVD manipulaciju klasifikatora imports", "Datu apmaijna"),
    ]),
    
    # 39. Bridinajumi un pazinojumi
    ("Bridinajumi un pazinojumi", [
        ("Kļudu sutišana uz e-pastu", "Bridinajumi un pazinojumi"),
        ("Kļudu pieteikumu registrēsana", "Bridinajumi un pazinojumi"),
        ("Kļudu ziņapmaijnas grupas iestatīsana", "Bridinajumi un pazinojumi"),
    ]),
]

# Flatten tasks and write to Excel
current_row = 3
for category_name, task_list in tasks:
    # Add category header
    ws.merge_cells(f'A{current_row}:F{current_row}')
    category_cell = ws.cell(row=current_row, column=1, value=category_name)
    category_cell.font = Font(bold=True, size=11)
    category_cell.fill = category_fill
    category_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[current_row].height = 22
    current_row += 1
    
    # Add tasks
    for task, cat in task_list:
        # Checkbox column (empty, user will fill)
        checkbox_cell = ws.cell(row=current_row, column=1, value="")
        checkbox_cell.alignment = Alignment(horizontal='center', vertical='center')
        checkbox_cell.border = border
        
        # Task name
        task_cell = ws.cell(row=current_row, column=2, value=task)
        task_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        task_cell.border = border
        
        # Category
        cat_cell = ws.cell(row=current_row, column=3, value=cat)
        cat_cell.alignment = Alignment(horizontal='left', vertical='center')
        cat_cell.border = border
        
        # Status dropdown cell (empty)
        status_cell = ws.cell(row=current_row, column=4, value="Nav sakts")
        status_cell.alignment = Alignment(horizontal='center', vertical='center')
        status_cell.border = border
        
        # Notes cell (empty)
        notes_cell = ws.cell(row=current_row, column=5, value="")
        notes_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        notes_cell.border = border
        
        # Date completed cell (empty)
        date_cell = ws.cell(row=current_row, column=6, value="")
        date_cell.alignment = Alignment(horizontal='center', vertical='center')
        date_cell.border = border
        
        ws.row_dimensions[current_row].height = 18
        current_row += 1

# Add data validation for Status column (column D)
dv = DataValidation(
    type="list",
    formula1='"Nav sakts;Macijos;Apguvu"',
    allow_blank=True
)
dv.error = 'Ludzu izvelieties no saraksta'
dv.errorTitle = 'Nepareiza vertiba'
dv.prompt = 'Izvelieties statusu'
dv.promptTitle = 'Statuss'

# Add validation to all data rows
first_data_row = 3
last_data_row = current_row - 1
ws.add_data_validation(dv)
dv.add(f'D{first_data_row}:D{last_data_row}')

# Freeze top rows
ws.freeze_panes = 'A3'

# Save the workbook
output_path = r'G:\Github\SS-WEB-SCRAPPER\Arsta_Birojs_Learning_Checklist.xlsx'
wb.save(output_path)
print(f"Excel fails veiksmigi izveidots: {output_path}")
print(f"Kopejais uzdevumu skaits: {sum(len(tasks_list) for _, tasks_list in tasks)}")
